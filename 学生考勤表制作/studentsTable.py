import openpyxl
import xlrd
import os
import shutil
print('从教务系统中学生注册模块下载学生注册表')
print('正在获取学生名单...')
wb = openpyxl.load_workbook('学生考勤表模板.xlsx')
S_moban = wb['模板']
workbook = xlrd.open_workbook('学生注册.xls')
S_stu = workbook.sheet_by_index(0)
# 创建以班级开头的三维数组
Students = []  # 【班级，【（学号1，姓名1，性别1）（学号2，姓名2，性别2），，，】】
for student in range(1, S_stu.nrows):
    group = S_stu.cell_value(student, 5)
    ID = S_stu.cell_value(student, 2)
    name = S_stu.cell_value(student, 1)
    gender = S_stu.cell_value(student, 3)
    student_info = (ID, name, gender)
    if not group:
        continue
    if '扩招' in group or '民航' in group:
        continue
    # 查找是否已存在该班级
    class_exists = False
    for class_item in Students:
        if class_item[0] == group:
            class_item[1].append(student_info)
            class_exists = True
            break
    # 如果班级不存在，则创建新班级
    if not class_exists:
        class_item = (group, [student_info])
        Students.append(class_item)
print('已获取学生名单')
print('正在生成考勤表总表...')
genderSet = set()
for row in Students:
    group = row[0]
    students = row[1]
    # 获取年级分组
    group_name = group[:5]
    genderSet.add(group_name)
    # 复制模板表并重命名为考场号
    copied_sheet = wb.copy_worksheet(S_moban)
    copied_sheet.title = group
    copied_sheet['A2'].value = '班级：' + group
    # 创建学号对应的单元格对象的一维数组
    tableId = [cell for col_range in ['B6:B25', 'B29:B53', 'B56:B100'] for cell in wb[group][col_range]]
    # 创建姓名对应的单元格对象的一维数组
    tableName = [cell for col_range in ['C6:C25', 'C29:C53', 'C56:C100'] for cell in wb[group][col_range]]
    # 创建性别对应的单元格对象的一维数组
    tableGender = [cell for col_range in ['D6:D25', 'D29:D53', 'D56:D100'] for cell in wb[group][col_range]]
    for index, student in enumerate(students):
        tableName[index][0].value = student[1]
        tableId[index][0].value = student[0]
        tableGender[index][0].value = student[2]
wb.remove(S_moban)
wb.save('考勤表总表.xlsx')
wb.close()
print('已生成考勤表总表')
folderName = '各年级考勤表'
if not os.path.exists(folderName):
    os.makedirs(folderName)
print('已生成分年级的考勤表的文件夹')
print('正在生成分年级的考勤表...')
for genderGroup in genderSet:
    wbSub = openpyxl.load_workbook('考勤表总表.xlsx')
    for sheet in wbSub.worksheets:
        if genderGroup not in sheet.title:
            wbSub.remove(sheet)
    # wbSub.create_sheet("New Sheet")
    name = f'{genderGroup}考勤表.xlsx'
    wbSub.save(name)
    wbSub.close()
    shutil.move(name, os.path.join(folderName, name))
print('完成操作！')
